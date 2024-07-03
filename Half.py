import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from email.message import EmailMessage
import smtplib
import config

# Paths
download_directory = os.path.join(os.getcwd(), "downloads")
os.makedirs(download_directory, exist_ok=True)

# Microsoft credentials
microsoft_email = config.SENDER_EMAIL
microsoft_password = config.SENDER_PASSWORD

# SMTP email configuration
smtp_server = config.SMTP_SERVER
smtp_port = config.SMTP_PORT
smtp_username = config.SENDER_EMAIL
smtp_password = config.SENDER_PASSWORD

# Email details
email_to = ['yash.srivastava@fareportal.com']
email_cc = ['palak.verma@fareportal.com']

def apply_styles(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

def combine_csv_files_to_excel(csv_files, output_excel_path):
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        for csv_file in csv_files:
            df = pd.read_csv(csv_file)
            sheet_name = os.path.basename(csv_file).replace('.csv', '')
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            apply_styles(sheet)

def filter_zero_hit_count_rows(all_data_df):
    zero_hit_count_df = all_data_df[all_data_df['Hit Count'] == '0']
    return zero_hit_count_df

def send_email_with_attachment(smtp_server, smtp_port, smtp_username, smtp_password, email_to, email_cc, subject, body, attachment_path):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_username
    msg['To'] = ', '.join(email_to)
    msg['Cc'] = ', '.join(email_cc)
    msg.set_content(body)

    with open(attachment_path, 'rb') as file:
        file_data = file.read()
        file_name = os.path.basename(file.name)
        msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)

    print(f"Email sent with attachment {attachment_path}")

def process_firewall_policies(csv_path, location):
    df = pd.read_csv(csv_path)

    active_policies_df = df[df['Status'] == 'Enabled']
    disabled_policies_df = df[df['Status'] == 'Disabled']
    zero_hit_count_df = df[df['Hit Count'] == '0']
    last_used_monthwise_df = df[df['Last Used'].notnull()].sort_values(by='Last Used', ascending=False)
    sorted_by_hit_count_df = df.sort_values(by='Hit Count', ascending=True)  # Sort in ascending order
    source_dest_all_df = df[(df['Source Address'] == 'all') & (df['Destination Address'] == 'all')]

    excel_path = os.path.join(download_directory, f'firewall_policies_{location}.xlsx')

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        active_policies_df.to_excel(writer, sheet_name='Active Policies', index=False)
        disabled_policies_df.to_excel(writer, sheet_name='Disabled Policies', index=False)
        zero_hit_count_df.to_excel(writer, sheet_name='0 Hit Count', index=False)
        last_used_monthwise_df.to_excel(writer, sheet_name='Last Used Month Wise', index=False)
        sorted_by_hit_count_df.to_excel(writer, sheet_name='Sorted by Hit Count', index=False)
        source_dest_all_df.to_excel(writer, sheet_name='Source Dest All', index=False)
        df.to_excel(writer, sheet_name='All Data', index=False)

    workbook = load_workbook(excel_path)
    for sheet_name in ['Active Policies', 'Disabled Policies', '0 Hit Count', 'Last Used Month Wise', 'Sorted by Hit Count', 'Source Dest All', 'All Data']:
        sheet = workbook[sheet_name]
        apply_styles(sheet)

    dashboard_sheet = workbook.create_sheet(title='Dashboard')

    def add_table(sheet, df, title, start_row, start_col, color):
        title_cell = sheet.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        headers = ['Name', 'First Used', 'Hit Count', 'ID', 'Last Used', 'Packets', 'Status', 'Comments']
        for col_idx, header in enumerate(headers, start_col):
            header_cell = sheet.cell(row=start_row + 1, column=col_idx, value=header)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row_idx, row in enumerate(dataframe_to_rows(df[headers], index=False, header=False), start_row + 2):
            for col_idx, value in enumerate(row, start_col):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    add_table(dashboard_sheet, active_policies_df, "Active Policies", 1, 1, "FFC7CE")
    add_table(dashboard_sheet, disabled_policies_df, "Disabled Policies", 1, 14, "FFEB9C")
    add_table(dashboard_sheet, zero_hit_count_df, "0 Hit Count", 1, 25, "C6EFCE")
    add_table(dashboard_sheet, last_used_monthwise_df, "Last Used Month Wise", 1, 36, "9BC2E6")
    add_table(dashboard_sheet, source_dest_all_df, "Source Dest All", 1, 47, "D9EAD3")  # New table for Source Dest All

    apply_styles(dashboard_sheet)

    # Highlight rows in "Last Used Month Wise" sheet
    last_used_sheet = workbook['Last Used Month Wise']
    for row in last_used_sheet.iter_rows(min_row=2, max_row=last_used_sheet.max_row, min_col=1, max_col=last_used_sheet.max_column):
        if row[4].value:  # Check if "Last Used" column has a value
            for cell in row:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    workbook.save(excel_path)

    email_subject = f"Firewall Policies for {location}"
    email_body = f'''Hi Team,

Please find the attached firewall policies of {location}.

Thanks & Regards,
EMS Team
'''

    send_email_with_attachment(
        smtp_server, smtp_port, smtp_username, smtp_password, email_to, email_cc, email_subject, email_body, excel_path
    )

    os.remove(excel_path)
    os.remove(csv_path)

# Define the CSV file paths and corresponding locations
csv_files = [
    ('path_to_your_csv_file_1.csv', 'Location 1'),
    # Add more tuples as needed
]

for csv_path, location in csv_files:
    process_firewall_policies(csv_path, location)

print("Script executed successfully.")
