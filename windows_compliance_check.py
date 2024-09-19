import subprocess
import openpyxl
from openpyxl import Workbook


def run_powershell_command(command):
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    return stdout.decode().strip(), stderr.decode().strip()


def check_compliance_windows(ip_address):
    compliance = {}

    # Check OS Version (End-of-Life status)
    os_command = f'powershell (Get-WmiObject -Class Win32_OperatingSystem).Caption'
    os_info, _ = run_powershell_command(os_command)
    if 'Windows Server 2016' in os_info or 'Windows Server 2019' in os_info or 'Windows Server 2022' in os_info:
        compliance['OS'] = "Non-EOL OS"
    else:
        compliance['OS'] = "EOL OS"

    # Check if SentinelOne Antivirus is installed and running
    sentinel_command = f'powershell Get-Service -Name "SentinelAgent"'
    sentinel_status, _ = run_powershell_command(sentinel_command)
    if 'Running' in sentinel_status:
        compliance['SentinelOne'] = "Installed and Running"
    else:
        compliance['SentinelOne'] = "Not Installed or Not Running"

    # Check if logging is enabled
    logging_command = f'powershell wevtutil qe Security /rd:true /c:1'
    log_status, _ = run_powershell_command(logging_command)
    if 'EventRecordID' in log_status:
        compliance['Logging'] = "Enabled"
    else:
        compliance['Logging'] = "Not Enabled"

    # Check if security patches are up to date
    patches_command = f'powershell Get-HotFix'
    patch_status, _ = run_powershell_command(patches_command)
    if patch_status:
        compliance['Patches'] = "Up to date"
    else:
        compliance['Patches'] = "Not Up to Date"

    # Check NTP configuration
    ntp_command = f'powershell w32tm /query /status'
    ntp_status, _ = run_powershell_command(ntp_command)
    if 'NtpServer' in ntp_status:
        compliance['NTP'] = ntp_status.split('NtpServer: ')[1].split('\n')[0]
    else:
        compliance['NTP'] = "Not Configured"

    # Check if necessary services are running (e.g., WinCollect for logging)
    wincollect_command = f'powershell Get-Service -Name "WinCollect"'
    wincollect_status, _ = run_powershell_command(wincollect_command)
    if 'Running' in wincollect_status:
        compliance['Services'] = "WinCollect Running"
    else:
        compliance['Services'] = "WinCollect Not Running"

    return compliance


def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    headers = ['OS', 'SentinelOne', 'Logging', 'Patches', 'NTP', 'Services']
    for idx, header in enumerate(headers, start=3):  
        sheet.cell(row=1, column=idx).value = header
    return wb, sheet


def main():
    file_path = 'servers_list.xlsx'
    wb, sheet = load_excel(file_path)

    
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=False):
        asset_name = row[0].value  
        asset_ip = row[1].value    
        print(f"Checking {asset_name} ({asset_ip})...")

        
        compliance_results = check_compliance_windows(asset_ip)

       
        for idx, (key, result) in enumerate(compliance_results.items(), start=2):  # Write starting from column C
            sheet.cell(row=row[0].row, column=idx + 1).value = result

    
    wb.save(file_path)
    print("Compliance report generated successfully.")

if __name__ == '__main__':
    main()
