import os
import shutil
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from email.message import EmailMessage
import smtplib
import config

# Configure WebDriver options
download_directory = os.path.join(os.getcwd(), "downloads")
temp_directory = os.path.join(os.getcwd(), "temp")
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": temp_directory,
    "download.prompt_for_download": False,
    "safebrowsing.enabled": True
})
driver_path = 'C:\\Users\\y.s.va22\\Downloads\\prod\\chromedriver.exe'

# List of FortiGate firewall IPs
firewall_ips = [
    '172.16.125.10',
]

# Shared FortiGate credentials
username_str = config.FW_USERNAME
password_str = config.FW_PASSWORD

# Microsoft credentials
microsoft_email = config.SENDER_EMAIL
microsoft_password = config.SENDER_PASSWORD

# SMTP email configuration
smtp_server = config.SMTP_SERVER
smtp_port = config.SMTP_PORT
smtp_username = config.SENDER_EMAIL
smtp_password = config.SENDER_PASSWORD

# Email details
email_to = ['yash.srivastava@fareportal.com', 'palak.verma@fareportal.com']
email_cc = ['akshat.jain@fareportal.com']
email_subject = 'Firewall Policy Review & Update'
email_body = '''Hi Team,

Please find the attached firewall policies of NJDC Lab Firewall.

Thanks & Regards,
EMS Team
'''

# Function to sign in to Microsoft
def sign_in(driver, email, password):
    EMAILFIELD = (By.ID, "i0116")
    PASSWORDFIELD = (By.ID, "i0118")
    NEXTBUTTON = (By.ID, "idSIButton9")

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(EMAILFIELD)).send_keys(email)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(NEXTBUTTON)).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(PASSWORDFIELD)).send_keys(password)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(NEXTBUTTON)).click()
    try:
        stay_signed_in_popup = driver.find_element(By.XPATH, "//div[contains(text(), 'Stay signed in?')]")
        if stay_signed_in_popup:
            no_button = driver.find_element(By.XPATH, "//input[@value='No']")
            no_button.click()
    except:
        pass

# Function to wait for file to be fully downloaded and rename it
def wait_for_download_and_rename(temp_dir, new_name, timeout=30):
    start_time = time.time()
    while True:
        for filename in os.listdir(temp_dir):
            if filename.endswith(".csv") and not filename.endswith(".crdownload"):
                full_path = os.path.join(temp_dir, filename)
                new_path = os.path.join(download_directory, new_name)
                shutil.move(full_path, new_path)
                print(f"File renamed and moved to {new_path}")
                return
        if time.time() - start_time > timeout:
            raise Exception(f"File download timed out after {timeout} seconds.")
        time.sleep(1)

# Function to export firewall policies
def export_policies(ip):
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get('https://login.microsoftonline.com/')
        sign_in(driver, microsoft_email, microsoft_password)
        driver.get(f'https://{ip}:8443/')
        try:
            advanced_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "details-button")))
            advanced_button.click()
            proceed_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "proceed-link")))
            proceed_link.click()
        except Exception as e:
            print(f"No advanced option found or proceed (unsafe) link: {e}")
        try:
            accept_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@name='accept']")))
            accept_button.click()
        except Exception as e:
            print(f"No accept button found: {e}")

        time.sleep(3)

        username = driver.find_element(By.ID, 'username')
        password = driver.find_element(By.ID, 'secretkey')
        login_button = driver.find_element(By.ID, 'login_button')

        username.send_keys(username_str)
        password.send_keys(password_str)
        login_button.click()

        time.sleep(5)

        driver.get(f'https://{ip}:8443/ng/firewall/policy/policy/standard')
        time.sleep(30)

        try:
            configure_table_button = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//f-icon[@class="fa-cog"]'))
            )
            driver.execute_script("arguments[0].click();", configure_table_button)
            print("Clicked on configure table button using JavaScript")
        except Exception as e:
            print(f"Failed to click configure table button: {e}")
            driver.save_screenshot('error_screenshot.png')
            return

        options = [
            "Destination Address",
            "First Used",
            "Hit Count",
            "ID",
            "IPS",
            "Last Used",
            "Packets",
            "Source Address",
            "Status",
            "Comments"
        ]

        for option in options:
            try:
                button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f'//button[.//span[text()="{option}"]]'))
                )
                driver.execute_script("arguments[0].click();", button)
                print(f"Clicked on '{option}' button")
            except Exception as e:
                print(f"Failed to click on '{option}' button: {e}")

        try:
            ok_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[@class="btn btn-primary" and text()="OK"]'))
            )
            driver.execute_script("arguments[0].click();", ok_button)
            print("Clicked on OK button to apply changes")
        except Exception as e:
            print(f"Failed to click OK button: {e}")
            return

        time.sleep(5)

        export_button = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//f-icon[@class="fa-download"]'))
        )
        driver.execute_script("arguments[0].click();", export_button)
        print("Clicked on export button using JavaScript")

        wait_for_download_and_rename(temp_directory, f"{ip}_policies.csv")

    finally:
        driver.quit()

# Process the downloaded files and update the Excel sheet
for ip in firewall_ips:
    export_policies(ip)

excel_path = 'path/to/your/excel/file.xlsx'
workbook = load_workbook(excel_path)
dashboard_sheet = workbook['Dashboard']

active_policies_df = pd.read_csv(os.path.join(download_directory, f"{firewall_ips[0]}_policies.csv"))
blocked_policies_df = pd.read_csv(os.path.join(download_directory, f"{firewall_ips[0]}_policies.csv"))
zero_hit_count_df = pd.read_csv(os.path.join(download_directory, f"{firewall_ips[0]}_policies.csv"))
last_used_monthwise_df = pd.read_csv(os.path.join(download_directory, f"{firewall_ips[0]}_policies.csv"))

def add_table(sheet, df, title, start_row, start_col, color):
    # Adding title
    title_cell = sheet.cell(row=start_row, column=start_col, value=title)
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 6)

    # Adding headers
    headers = ['Name', 'First Used', 'Hit Count', 'ID', 'Last Used', 'Packets', 'Status']
    for col_idx, header in enumerate(headers, start=start_col):
        header_cell = sheet.cell(row=start_row + 1, column=col_idx, value=header)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_cell.alignment = Alignment(horizontal="center")

    # Adding data
    for row_idx, row in enumerate(dataframe_to_rows(df[headers], index=False, header=False), start=start_row + 2):
        for col_idx, value in enumerate(row, start=start_col):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Adding tables to dashboard
add_table(dashboard_sheet, active_policies_df, "Active Policies", 1, 2, "4F81BD")
add_table(dashboard_sheet, blocked_policies_df, "Blocked Policies", 1, 11, "4F81BD")
add_table(dashboard_sheet, zero_hit_count_df, "0 Hit Count", 1, 20, "4F81BD")
add_table(dashboard_sheet, last_used_monthwise_df, "Last Used Month Wise", 1, 29, "4F81BD")

workbook.save(excel_path)

# Function to send email with attachment
def send_email(subject, body, to, cc, attachment_path):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_username
    msg['To'] = ', '.join(to)
    msg['Cc'] = ', '.join(cc)
    msg.set_content(body)

    with open(attachment_path, 'rb') as f:
        file_data = f.read()
        file_name = os.path.basename(attachment_path)
        msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)

# Send email with updated Excel file as attachment
send_email(email_subject, email_body, email_to, email_cc, excel_path)
