import os
import shutil
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from email.message import EmailMessage
import smtplib
import config

# Configure WebDriver options
download_directory = os.path.join(os.getcwd(), "downloads")
temp_directory = os.path.join(os.getcwd(), "temp")
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": temp_directory,
    "download.prompt_for_download": False,
    "safebrowsing.enabled": True
})
driver_path = 'C:\\Users\\y.s.va22\\Downloads\\prod\\chromedriver.exe'

# List of FortiGate firewall IPs with location
firewall_ips = [
    ('172.16.125.10', 'OKR 1'),
    # ('10.255.255.65', 'NJDC LAB'),
    # ('172.16.125.12', 'Pune'),
    # ('172.16.125.13', 'Mexico'),
    # ('172.16.125.14', 'LI'),
    # ('172.16.125.15', 'Canada'),
    # ('172.16.125.16', 'New York'),
    # ('172.16.125.17', 'NDJC 1'),
    # ('172.16.125.18', 'LV'),
    # ('172.16.125.19', 'OKR 2'),
    # ('172.16.125.20', 'NDJC 2'),
    # ('172.16.125.21', 'NDJC 3')
]

# Shared FortiGate credentials
username_str = config.FW_USERNAME
password_str = config.FW_PASSWORD

# Microsoft credentials
microsoft_email = config.SENDER_EMAIL
microsoft_password = config.SENDER_PASSWORD

# SMTP email configuration
smtp_server = config.SMTP_SERVER
smtp_port = config.SMTP_PORT
smtp_username = config.SENDER_EMAIL
smtp_password = config.SENDER_PASSWORD

# Email details
email_to = ['yash.srivastava@fareportal.com']
email_cc = ['palak.verma@fareportal.com']

# Function to sign in to Microsoft
def sign_in(driver, email, password):
    EMAILFIELD = (By.ID, "i0116")
    PASSWORDFIELD = (By.ID, "i0118")
    NEXTBUTTON = (By.ID, "idSIButton9")

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(EMAILFIELD)).send_keys(email)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(NEXTBUTTON)).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(PASSWORDFIELD)).send_keys(password)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(NEXTBUTTON)).click()
    try:
        stay_signed_in_popup = driver.find_element(By.XPATH, "//div[contains(text(), 'Stay signed in?')]")
        if stay_signed_in_popup:
            no_button = driver.find_element(By.XPATH, "//input[@value='No']")
            no_button.click()
    except:
        pass

# Function to wait for file to be fully downloaded and rename it
def wait_for_download_and_rename(temp_dir, new_name, timeout=30):
    start_time = time.time()
    while True:
        for filename in os.listdir(temp_dir):
            if filename.endswith(".csv") and not filename.endswith(".crdownload"):
                full_path = os.path.join(temp_dir, filename)
                new_path = os.path.join(download_directory, new_name)
                shutil.move(full_path, new_path)
                print(f"File renamed and moved to {new_path}")
                return
        if time.time() - start_time > timeout:
            raise Exception(f"File download timed out after {timeout} seconds.")
        time.sleep(1)

# Function to export firewall policies
def export_policies(ip, location):
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get('https://login.microsoftonline.com/')
        sign_in(driver, microsoft_email, microsoft_password)
        driver.get(f'https://{ip}:8443/')
        try:
            advanced_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "details-button")))
            advanced_button.click()
            proceed_link = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "proceed-link")))
            proceed_link.click()
        except Exception as e:
            print(f"No advanced option found or proceed (unsafe) link: {e}")
        try:
            accept_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@name='accept']")))
            accept_button.click()
        except Exception as e:
            print(f"No accept button found: {e}")

        time.sleep(3)

        username = driver.find_element(By.ID, 'username')
        password = driver.find_element(By.ID, 'secretkey')
        login_button = driver.find_element(By.ID, 'login_button')

        username.send_keys(username_str)
        password.send_keys(password_str)
        login_button.click()

        time.sleep(5)

        # Handle special case for '172.16.125.20' with ?vdom=root
        if ip == '10.255.255.65':
            driver.get(f'https://{ip}:8443/ng/firewall/policy/policy/standard?vdom=root')
        else:
            driver.get(f'https://{ip}:8443/ng/firewall/policy/policy/standard')

        time.sleep(30)

        try:
            configure_table_button = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//f-icon[@class="fa-cog"]'))
            )
            driver.execute_script("arguments[0].click();", configure_table_button)
            print("Clicked on configure table button using JavaScript")
        except Exception as e:
            print(f"Failed to click configure table button: {e}")
            driver.save_screenshot('error_screenshot.png')
            return

        options = [
            "Destination Address",
            "First Used",
            "Hit Count",
            "ID",
            "IPS",
            "Last Used",
            "Packets",
            "Source Address",
            "Status",
            "Comments"
        ]

        for option in options:
            try:
                button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f'//button[.//span[text()="{option}"]]'))
                )
                driver.execute_script("arguments[0].click();", button)
                print(f"Clicked on '{option}' button")
            except Exception as e:
                print(f"Failed to click on '{option}' button: {e}")

        try:
            apply_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, "standard-button primary") and text()="Apply"]'))
            )
            driver.execute_script("arguments[0].click();", apply_button)
            print("Clicked on Apply button")
        except Exception as e:
            print(f"Failed to click on Apply button: {e}")

        try:
            export_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//span[@class="ng-binding ng-scope" and text()="Export"]'))
            )
            driver.execute_script("arguments[0].click();", export_button)
            time.sleep(2)
            print("Clicked on Export button")
        except Exception as e:
            print(f"Failed to click on Export button: {e}")

        try:
            export_csv_option = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//span[@class="ng-scope" and text()="CSV"]'))
            )
            driver.execute_script("arguments[0].click();", export_csv_option)
            print("Clicked on CSV option")
        except Exception as e:
            print(f"Failed to click on CSV option: {e}")

        new_filename = f'firewall_policies_{location}.csv'
        time.sleep(5)
        wait_for_download_and_rename(temp_directory, new_filename)

    except Exception as e:
        print(f"An error occurred while exporting policies for {ip}: {e}")
        driver.save_screenshot('error_screenshot.png')
    finally:
        driver.quit()

os.makedirs(download_directory, exist_ok=True)
os.makedirs(temp_directory, exist_ok=True)

def apply_styles(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

def combine_csv_files_to_excel(csv_files, output_excel_path):
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        for csv_file in csv_files:
            df = pd.read_csv(csv_file)
            sheet_name = os.path.basename(csv_file).replace('.csv', '')
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            apply_styles(sheet)

def filter_zero_hit_count_rows(all_data_df):
    zero_hit_count_df = all_data_df[all_data_df['Hit Count'] == '0']
    return zero_hit_count_df

def send_email_with_attachment(smtp_server, smtp_port, smtp_username, smtp_password, email_to, email_cc, subject, body, attachment_path):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = smtp_username
    msg['To'] = ', '.join(email_to)
    msg['Cc'] = ', '.join(email_cc)
    msg.set_content(body)

    with open(attachment_path, 'rb') as file:
        file_data = file.read()
        file_name = os.path.basename(file.name)
        msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(msg)

    print(f"Email sent with attachment {attachment_path}")

def process_firewall_policies(firewall_ip, location):
    export_policies(firewall_ip, location)

    csv_path = os.path.join(download_directory, f'firewall_policies_{location}.csv')

    df = pd.read_csv(csv_path)

    active_policies_df = df[df['Status'] == 'Enabled']
    disabled_policies_df = df[df['Status'] == 'Disabled']
    zero_hit_count_df = df[df['Hit Count'] == '0']
    last_used_monthwise_df = df[df['Last Used'].notnull()].sort_values(by='Last Used', ascending=False)
    sorted_by_hit_count_df = df.sort_values(by='Hit Count', ascending=True)  # Sort in ascending order
    source_dest_all_df = df[(df['Source Address'] == 'all') & (df['Destination Address'] == 'all')]

    excel_path = os.path.join(download_directory, f'firewall_policies_{location}.xlsx')

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        active_policies_df.to_excel(writer, sheet_name='Active Policies', index=False)
        disabled_policies_df.to_excel(writer, sheet_name='Disabled Policies', index=False)
        zero_hit_count_df.to_excel(writer, sheet_name='0 Hit Count', index=False)
        last_used_monthwise_df.to_excel(writer, sheet_name='Last Used Month Wise', index=False)
        sorted_by_hit_count_df.to_excel(writer, sheet_name='Sorted by Hit Count', index=False)
        source_dest_all_df.to_excel(writer, sheet_name='Source Dest All', index=False)
        df.to_excel(writer, sheet_name='All Data', index=False)

    workbook = load_workbook(excel_path)
    for sheet_name in ['Active Policies', 'Disabled Policies', '0 Hit Count', 'Last Used Month Wise', 'Sorted by Hit Count', 'Source Dest All', 'All Data']:
        sheet = workbook[sheet_name]
        apply_styles(sheet)

    dashboard_sheet = workbook.create_sheet(title='Dashboard')

    def add_table(sheet, df, title, start_row, start_col, color):
        title_cell = sheet.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True)
        title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        headers = ['Name', 'First Used', 'Hit Count', 'ID', 'Last Used', 'Packets', 'Status', 'Comments']
        for col_idx, header in enumerate(headers, start_col):
            header_cell = sheet.cell(row=start_row + 1, column=col_idx, value=header)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row_idx, row in enumerate(dataframe_to_rows(df[headers], index=False, header=False), start_row + 2):
            for col_idx, value in enumerate(row, start_col):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    add_table(dashboard_sheet, active_policies_df, "Active Policies", 1, 1, "FFC7CE")
    add_table(dashboard_sheet, disabled_policies_df, "Disabled Policies", 1, 14, "FFEB9C")
    add_table(dashboard_sheet, zero_hit_count_df, "0 Hit Count", 1, 25, "C6EFCE")
    add_table(dashboard_sheet, last_used_monthwise_df, "Last Used Month Wise", 1, 36, "9BC2E6")
    add_table(dashboard_sheet, source_dest_all_df, "Source Dest All", 1, 47, "D9EAD3")  # New table for Source Dest All

    apply_styles(dashboard_sheet)

    # Highlight rows in "Last Used Month Wise" sheet
    last_used_sheet = workbook['Last Used Month Wise']
    for row in last_used_sheet.iter_rows(min_row=2, max_row=last_used_sheet.max_row, min_col=1, max_col=last_used_sheet.max_column):
        if row[4].value:  # Check if "Last Used" column has a value
            for cell in row:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    workbook.save(excel_path)

    email_subject = f"Firewall Policies for {location}"
    email_body = f'''Hi Team,

Please find the attached firewall policies of {firewall_ip} ({location}).

Thanks & Regards,
EMS Team
'''

    send_email_with_attachment(
        smtp_server, smtp_port, smtp_username, smtp_password, email_to, email_cc, email_subject, email_body, excel_path
    )

    os.remove(excel_path)
    os.remove(csv_path)

for firewall_ip, location in firewall_ips:
    process_firewall_policies(firewall_ip, location)

print("Script executed successfully.")
    
